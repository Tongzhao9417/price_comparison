#!/usr/bin/env python3
"""Sync relay comparison table from xlsx/csv into docs/index.md."""

from __future__ import annotations

import csv
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

REPO_ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = REPO_ROOT / 'data' / 'relay_table.xlsx'
CSV_PATH = REPO_ROOT / 'data' / 'relay_table.csv'
DOC_PATH = REPO_ROOT / 'docs' / 'index.md'
START_MARKER = '<!-- relay-table:start -->'
END_MARKER = '<!-- relay-table:end -->'


def fail(message: str) -> None:
    raise SystemExit(message)


def normalize_float(value: float) -> str:
    # Avoid floating-point noise like 13.750000000000002 from formula cells.
    text = format(value, '.15g')
    if 'e' in text or 'E' in text:
        text = format(value, '.15f').rstrip('0').rstrip('.')
    if text == '-0':
        return '0'
    return text


def normalize_cell(value: object) -> str:
    if value is None:
        return ''
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return normalize_float(value)
    return str(value).strip()


def escape_cell(value: str) -> str:
    return value.replace('|', '\\|')


def trim_trailing_empty(row: list[str]) -> list[str]:
    last_index = len(row)
    while last_index > 0 and row[last_index - 1] == '':
        last_index -= 1
    return row[:last_index]


def parse_matrix(matrix: list[list[object]], source_name: str) -> tuple[list[str], list[list[str]]]:
    if not matrix:
        fail(f'{source_name} has no rows.')

    normalized_rows = [trim_trailing_empty([normalize_cell(cell) for cell in row]) for row in matrix]
    normalized_rows = [row for row in normalized_rows if any(cell != '' for cell in row)]

    if not normalized_rows:
        fail(f'{source_name} has no non-empty rows.')

    header = normalized_rows[0]
    if not header:
        fail(f'{source_name} header row is empty.')

    if any(cell == '' for cell in header):
        fail(f'{source_name} header contains empty column names.')

    header_column_count = len(header)

    data_rows: list[list[str]] = []
    for row_number, row in enumerate(normalized_rows[1:], start=2):
        if len(row) > header_column_count:
            fail(
                f'{source_name} row {row_number} has {len(row)} columns, '
                f'but header has {header_column_count}. Please fill header cells for new columns.'
            )

        padded_row = row + [''] * (header_column_count - len(row))
        if all(cell == '' for cell in padded_row):
            continue

        if padded_row[0] == '':
            fail(f'{source_name} row {row_number}: first column cannot be empty.')

        data_rows.append(padded_row)

    if not data_rows:
        fail(f'{source_name} has no data rows.')

    return header, data_rows


def load_from_xlsx() -> tuple[list[str], list[list[str]]]:
    if not XLSX_PATH.exists():
        fail(f'XLSX file not found: {XLSX_PATH}')
    if load_workbook is None:
        fail('openpyxl is required to read xlsx. Install with `python3 -m pip install openpyxl`.')

    workbook = load_workbook(XLSX_PATH, data_only=True)
    sheet = workbook.active
    matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
    return parse_matrix(matrix, str(XLSX_PATH))


def load_from_csv() -> tuple[list[str], list[list[str]]]:
    if not CSV_PATH.exists():
        fail(f'CSV file not found: {CSV_PATH}')

    with CSV_PATH.open('r', encoding='utf-8-sig', newline='') as csv_file:
        reader = csv.reader(csv_file)
        matrix = [row for row in reader]

    return parse_matrix(matrix, str(CSV_PATH))


def load_table_data() -> tuple[list[str], list[list[str]], str]:
    if XLSX_PATH.exists():
        header, rows = load_from_xlsx()
        return header, rows, 'xlsx'
    if CSV_PATH.exists():
        header, rows = load_from_csv()
        return header, rows, 'csv'
    fail(f'No table source found. Expected either {XLSX_PATH} or {CSV_PATH}.')


def write_csv_snapshot(header: list[str], rows: list[list[str]]) -> None:
    with CSV_PATH.open('w', encoding='utf-8-sig', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(header)
        writer.writerows(rows)


def build_table_markdown(header: list[str], rows: list[list[str]]) -> str:
    header_line = '|' + '|'.join(escape_cell(cell) for cell in header) + '|'
    align_line = '|' + '|'.join(':----:' for _ in header) + '|'

    markdown_rows = [header_line, align_line]
    for row in rows:
        markdown_rows.append('|' + '|'.join(escape_cell(cell) for cell in row) + '|')

    return '\n'.join(markdown_rows)


def sync_docs(table_markdown: str) -> None:
    if not DOC_PATH.exists():
        fail(f'Document file not found: {DOC_PATH}')

    doc_content = DOC_PATH.read_text(encoding='utf-8')
    marker_pattern = re.compile(
        re.escape(START_MARKER) + r'.*?' + re.escape(END_MARKER),
        flags=re.S,
    )

    if not marker_pattern.search(doc_content):
        fail(
            f'Marker block not found in {DOC_PATH}. '
            f'Please add:\n{START_MARKER}\n...\n{END_MARKER}'
        )

    updated_block = f'{START_MARKER}\n{table_markdown}\n{END_MARKER}'
    updated_content = marker_pattern.sub(updated_block, doc_content, count=1)

    if updated_content == doc_content:
        print('No table changes detected.')
        return

    DOC_PATH.write_text(updated_content, encoding='utf-8')
    print('Synced table into docs/index.md')


def main() -> None:
    header, rows, source = load_table_data()
    if source == 'xlsx':
        write_csv_snapshot(header, rows)
    table_markdown = build_table_markdown(header, rows)
    sync_docs(table_markdown)


if __name__ == '__main__':
    main()
